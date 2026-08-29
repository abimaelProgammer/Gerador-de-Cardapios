"""Gera o cardapio a partir de uma planilha de produtos.

Uso pela linha de comando:
    python functions.py --produtos "Produtos (58).xlsx" --saida "Cardapio.xlsx"

A fonte pode ser trocada tanto pelo argumento ``--produtos`` quanto ao chamar
diretamente a funcao ``gerar_cardapio``.
"""

from __future__ import annotations

import argparse
from collections import OrderedDict
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


COLUNAS_OBRIGATORIAS = {
    "codigo": ("Cód. Sistema", "Cod. Sistema", "Código", "Codigo"),
    "categoria": ("Categoria",),
    "nome": ("Nome", "Produto", "Item"),
    "preco": ("Preço Venda", "Preco Venda", "Preço", "Preco"),
    "status": ("Status Venda", "Status"),
}

RODAPES_PADRAO = [
    ("Cobramos 10% como taxa de serviço", "Cobramos 10% como taxa de serviço"),
    (
        "Rede WI-FI: Clientes Le Clair    Senha: @Clientes2026",
        "Rede WI-FI: Clientes Le Clair    Senha: @Clientes2026",
    ),
    ("PIX (CNPJ) 64.111.665/0001-55", "PIX (CPF) 428.255.047-34"),
]


def _normalizar(texto: Any) -> str:
    import unicodedata

    texto = "" if texto is None else str(texto)
    return "".join(
        c for c in unicodedata.normalize("NFKD", texto) if not unicodedata.combining(c)
    ).strip().casefold()


def _localizar_colunas(cabecalhos: Iterable[Any]) -> dict[str, int]:
    normalizados = {_normalizar(valor): indice for indice, valor in enumerate(cabecalhos, 1)}
    resultado: dict[str, int] = {}
    for campo, alternativas in COLUNAS_OBRIGATORIAS.items():
        for alternativa in alternativas:
            if _normalizar(alternativa) in normalizados:
                resultado[campo] = normalizados[_normalizar(alternativa)]
                break
        if campo not in resultado and campo != "status":
            raise ValueError(
                f"Coluna obrigatória não encontrada: {alternativas[0]}. "
                "Verifique o cabeçalho da planilha de produtos."
            )
    return resultado


def ler_produtos(fonte_produtos: str | Path, aba: str | None = None) -> list[dict[str, Any]]:
    """Lê produtos de qualquer .xlsx com os cabeçalhos esperados."""
    caminho = Path(fonte_produtos)
    if not caminho.exists():
        raise FileNotFoundError(f"Fonte de produtos não encontrada: {caminho}")

    wb = load_workbook(caminho, data_only=True, read_only=True)
    try:
        ws = wb[aba] if aba else wb[wb.sheetnames[0]]
        colunas = _localizar_colunas(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        produtos = []
        for numero_linha, linha in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            def valor(campo: str) -> Any:
                indice = colunas.get(campo)
                return linha[indice - 1] if indice and indice <= len(linha) else None

            if valor("codigo") is None or not valor("nome"):
                continue
            produtos.append(
                {
                    "codigo": valor("codigo"),
                    "categoria": valor("categoria") or "SEM CATEGORIA",
                    "nome": valor("nome"),
                    "preco": valor("preco") or 0,
                    "status": valor("status") or "Ativo",
                    "linha_origem": numero_linha,
                }
            )
        return produtos
    finally:
        wb.close()


def _rodapes_do_modelo(
    modelo_cardapio: str | Path | None,
) -> list[tuple[str, str]]:
    """Aproveita os avisos das duas metades do cardápio anterior."""
    if not modelo_cardapio or not Path(modelo_cardapio).exists():
        return list(RODAPES_PADRAO)
    wb = load_workbook(modelo_cardapio, data_only=False, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        textos: list[tuple[str, str]] = []
        palavras_chave = (
            "taxa", "wi-fi", "wifi", "pix", "servico", "gorjeta",
            "funcionamento", "reserva", "pedido minimo", "senha", "cnpj", "cpf",
            "pagamento", "cartao", "dinheiro",
        )

        def texto_do_bloco(linha: tuple[Any, ...], inicio: int, fim: int) -> str:
            for valor in linha[inicio:fim]:
                if valor is None:
                    continue
                texto = str(valor).strip()
                if texto and not texto.startswith("="):
                    return texto
            return ""

        for linha in ws.iter_rows(values_only=True):
            # A coluna A guarda o marcador S/N. Os avisos ficam nos blocos
            # B:D (esquerda) e E:G (direita) do modelo original.
            texto_esquerda = texto_do_bloco(linha, 1, 4)
            texto_direita = texto_do_bloco(linha, 4, 7)
            candidatos = (texto_esquerda, texto_direita)

            if not any(
                any(chave in _normalizar(texto) for chave in palavras_chave)
                for texto in candidatos if texto
            ):
                continue

            texto_esquerda = texto_esquerda or texto_direita
            texto_direita = texto_direita or texto_esquerda
            rodape = (texto_esquerda, texto_direita)
            if max(map(len, rodape)) > 10 and rodape not in textos:
                textos.append(rodape)
        return textos or list(RODAPES_PADRAO)
    finally:
        wb.close()


def gerar_cardapio(
    fonte_produtos: str | Path,
    destino: str | Path = "Cardapio.xlsx",
    modelo_cardapio: str | Path | None = "cardapio 2026-07-28.xlsx",
    *,
    aba_produtos: str | None = None,
    ocultar_pausados: bool = True,
    categorias_excluidas: Iterable[str] = (),
) -> Path:
    """Cria um novo Excel de cardápio e retorna o caminho gerado.

    ``fonte_produtos`` é deliberadamente um parâmetro para permitir trocar o
    relatório de produtos sem alterar este arquivo.
    """
    produtos = ler_produtos(fonte_produtos, aba_produtos)
    excluidas = {_normalizar(c) for c in categorias_excluidas}
    produtos = [p for p in produtos if _normalizar(p["categoria"]) not in excluidas]

    grupos: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for produto in produtos:
        grupos.setdefault(str(produto["categoria"]), []).append(produto)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cardápio"
    apoio = wb.create_sheet("Configuração")

    azul = "1F4E78"
    azul_claro = "D9EAF7"
    branco = "FFFFFF"
    cinza = "D9E1F2"
    bege_creme = "EBE2D1"
    verde_escuro = "1B291E"
    borda_fina = Border(*(Side(style="thin", color="808080") for _ in range(4)))

    ws.merge_cells("A1:D1")
    ws["A1"] = f"CARDÁPIO {date.today():%d/%m/%Y}"
    ws["A1"].font = Font(size=18, bold=True, color=verde_escuro)
    ws["A1"].fill = PatternFill("solid", fgColor=bege_creme)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("F1:I1")
    ws["F1"] = f"CARDÁPIO {date.today():%d/%m/%Y}"
    ws["F1"].font = Font(size=18, bold=True, color=verde_escuro)
    ws["F1"].fill = PatternFill("solid", fgColor=bege_creme)
    ws["F1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[1].height = 60
    
    caminho_logo = Path(fonte_produtos).parent / "logo2.jpg"
    if caminho_logo.exists():
        img = Image(str(caminho_logo))
        ratio = 70 / img.height
        img.width = int(img.width * ratio)
        img.height = 70
        ws.add_image(img, "A1")
        
        img2 = Image(str(caminho_logo))
        img2.width = img.width
        img2.height = img.height
        ws.add_image(img2, "F1")

    cabecalhos = ("IMP", "CÓD", "ITENS", "R$")
    for coluna, texto in enumerate(cabecalhos, 1):
        celula = ws.cell(2, coluna, texto)
        celula.font = Font(bold=True, color=branco)
        celula.fill = PatternFill("solid", fgColor=azul)
        celula.alignment = Alignment(horizontal="center")
        celula.border = borda_fina

        celula_dir = ws.cell(2, coluna + 5, texto)
        celula_dir.font = Font(bold=True, color=branco)
        celula_dir.fill = PatternFill("solid", fgColor=azul)
        celula_dir.alignment = Alignment(horizontal="center")
        celula_dir.border = borda_fina

    linha_atual = 3
    ativos = pausados = 0
    for categoria, itens in grupos.items():
        ws.cell(linha_atual, 1, "S")
        ws.cell(linha_atual, 3, categoria.upper())
        ws.cell(linha_atual, 3).alignment = Alignment(horizontal="center")
        
        ws.cell(linha_atual, 6, "S")
        ws.cell(linha_atual, 8, categoria.upper())
        ws.cell(linha_atual, 8).alignment = Alignment(horizontal="center")
        
        for coluna in range(1, 5):
            celula = ws.cell(linha_atual, coluna)
            celula.fill = PatternFill("solid", fgColor=azul_claro)
            celula.font = Font(bold=True, color=azul)
            celula.border = borda_fina
            
            celula_dir = ws.cell(linha_atual, coluna + 5)
            celula_dir.fill = PatternFill("solid", fgColor=azul_claro)
            celula_dir.font = Font(bold=True, color=azul)
            celula_dir.border = borda_fina
            
        tem_ativo = any(_normalizar(produto["status"]) == "ativo" for produto in itens)
        if not tem_ativo and ocultar_pausados:
            ws.row_dimensions[linha_atual].hidden = True
            
        linha_atual += 1

        for produto in itens:
            ativo = _normalizar(produto["status"]) == "ativo"
            ws.cell(linha_atual, 1, "S" if ativo else "N")
            ws.cell(linha_atual, 2, produto["codigo"])
            ws.cell(linha_atual, 2).alignment = Alignment(horizontal="center")
            ws.cell(linha_atual, 3, produto["nome"])
            ws.cell(linha_atual, 4, produto["preco"])
            ws.cell(linha_atual, 4).number_format = 'R$ #,##0.00'
            
            ws.cell(linha_atual, 6, "S" if ativo else "N")
            ws.cell(linha_atual, 7, produto["codigo"])
            ws.cell(linha_atual, 7).alignment = Alignment(horizontal="center")
            ws.cell(linha_atual, 8, produto["nome"])
            ws.cell(linha_atual, 9, produto["preco"])
            ws.cell(linha_atual, 9).number_format = 'R$ #,##0.00'
            
            for coluna in range(1, 5):
                ws.cell(linha_atual, coluna).border = borda_fina
                ws.cell(linha_atual, coluna + 5).border = borda_fina
                
            if not ativo:
                pausados += 1
                ws.row_dimensions[linha_atual].hidden = ocultar_pausados
                for coluna in range(1, 5):
                    ws.cell(linha_atual, coluna).font = Font(color="808080", italic=True)
                    ws.cell(linha_atual, coluna + 5).font = Font(color="808080", italic=True)
            else:
                ativos += 1
            linha_atual += 1

    rodapes = _rodapes_do_modelo(modelo_cardapio)
    if rodapes:
        linha_atual += 1
        for texto_esquerda, texto_direita in rodapes:
            ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=4)
            celula = ws.cell(linha_atual, 1, texto_esquerda)
            celula.alignment = Alignment(horizontal="center")
            celula.font = Font(bold=True)
            celula.fill = PatternFill("solid", fgColor=cinza)
            
            ws.merge_cells(start_row=linha_atual, start_column=6, end_row=linha_atual, end_column=9)
            celula_dir = ws.cell(linha_atual, 6, texto_direita)
            celula_dir.alignment = Alignment(horizontal="center")
            celula_dir.font = Font(bold=True)
            celula_dir.fill = PatternFill("solid", fgColor=cinza)
            
            linha_atual += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 68
    ws.column_dimensions["D"].width = 15
    
    ws.column_dimensions["E"].width = 2
    
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["F"].hidden = True
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 68
    ws.column_dimensions["I"].width = 15
    
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:D{linha_atual - 1}"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:2"
    ws.print_area = f"A1:I{linha_atual - 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    configuracoes = (
        ("Configuração", "Valor"),
        ("Fonte de produtos", str(Path(fonte_produtos).resolve())),
        ("Modelo consultado", str(Path(modelo_cardapio).resolve()) if modelo_cardapio else ""),
        ("Produtos incluídos", len(produtos)),
        ("Produtos ativos", ativos),
        ("Produtos pausados", pausados),
        ("Pausados ocultos", "Sim" if ocultar_pausados else "Não"),
    )
    for linha in configuracoes:
        apoio.append(linha)
    for celula in apoio[1]:
        celula.font = Font(bold=True, color=branco)
        celula.fill = PatternFill("solid", fgColor=azul)
    apoio.column_dimensions["A"].width = 24
    apoio.column_dimensions["B"].width = 90

    caminho_destino = Path(destino)
    caminho_destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_destino)
    return caminho_destino.resolve()


def _argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera Cardapio.xlsx a partir de um Excel de produtos.")
    parser.add_argument("--produtos", default="Produtos (58).xlsx", help="Excel que será a fonte dos produtos")
    parser.add_argument("--modelo", default="cardapio 2026-07-28.xlsx", help="Cardápio anterior usado como referência")
    parser.add_argument("--saida", default="Cardapio.xlsx", help="Nome/caminho do novo Excel")
    parser.add_argument("--mostrar-pausados", action="store_true", help="Não oculta as linhas de produtos pausados")
    return parser.parse_args()


if __name__ == "__main__":
    args = _argumentos()
    gerado = gerar_cardapio(
        args.produtos,
        args.saida,
        args.modelo,
        ocultar_pausados=not args.mostrar_pausados,
    )
    print(f"Cardápio gerado: {gerado}")
